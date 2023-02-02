def load_from_xl(file_name):
    import numpy as np
    import openpyxl as xl
    print("Loading input training data from excel spreadsheet...")
    # data_only = True means the value is read not the formula
    wb = xl.load_workbook(file_name, data_only=True)
    sheet = wb["Sheet1"]
    x_train = np.zeros(sheet.max_row - 1, dtype=float)
    y_train = np.zeros(sheet.max_row - 1, dtype=float)
    for row in range(2, sheet.max_row + 1):
        x_train[row-2] = sheet.cell(row, 2).value
        y_train[row-2] = sheet.cell(row, 3).value
    return x_train, y_train